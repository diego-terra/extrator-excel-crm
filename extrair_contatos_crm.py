"""
extrair_contatos_crm.py
=======================
Varre uma pasta com arquivos .xlsx de pedidos,
extrai CLIENTE e CONTATO diretamente das células da aba Plan1,
remove duplicatas e gera CRM_contatos.xlsx.

PADRÃO DETECTADO NOS ARQUIVOS:
  Aba "Plan1", linha 3 (aprox):
    "CLIENTE:" → próxima célula = nome do cliente
    "CONTATO:" → próxima célula = telefone

REQUISITOS:
  pip install pandas openpyxl

USO:
  python extrair_contatos_crm.py
  python extrair_contatos_crm.py "C:/Pedidos"
"""

import os
import sys
import re
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

OUTPUT_FILE = "CRM_contatos.xlsx"


def listar_xlsx(pasta):
    arquivos = []
    for raiz, _, nomes in os.walk(pasta):
        for nome in nomes:
            if nome.lower().endswith(".xlsx") and not nome.startswith("~$"):
                arquivos.append(os.path.join(raiz, nome))
    return sorted(arquivos)


def extrair_contato(caminho):
    nome_arquivo = os.path.basename(caminho)
    resultado = {"nome": "", "telefone": "", "arquivo": nome_arquivo}

    try:
        wb = load_workbook(caminho, read_only=True, data_only=True)
    except Exception as e:
        resultado["erro"] = str(e)
        return resultado

    # Estratégia 1: aba Plan1 — busca rótulos CLIENTE: e CONTATO: na linha
    aba_nomes_principais = {"plan1", "plan 1", "pedido", "orçamento", "orcamento"}
    aba_principal = None
    for n in wb.sheetnames:
        if n.strip().lower() in aba_nomes_principais:
            aba_principal = wb[n]
            break
    if aba_principal is None and wb.sheetnames:
        aba_principal = wb[wb.sheetnames[0]]

    if aba_principal:
        for row in aba_principal.iter_rows(max_row=10, values_only=True):
            cells = [str(c).strip().upper() if c is not None else "" for c in row]
            for i, cell in enumerate(cells):
                if cell in ("CLIENTE:", "CLIENTE") and not resultado["nome"]:
                    for j in range(i + 1, len(cells)):
                        v = cells[j]
                        if v and v not in ("CLIENTE:", "CLIENTE", "CONTATO:", "CONTATO", "Nº", "N°", "DATA"):
                            resultado["nome"] = str(row[j]).strip()
                            break
                if cell in ("CONTATO:", "CONTATO") and not resultado["telefone"]:
                    for j in range(i + 1, len(cells)):
                        v = cells[j]
                        if v and v not in ("CLIENTE:", "CLIENTE", "CONTATO:", "CONTATO", "Nº", "N°", "DATA"):
                            resultado["telefone"] = str(row[j]).strip()
                            break
            if resultado["nome"]:
                break

    # Estratégia 2: aba Planilha1 com cabeçalho (fallback)
    if not resultado["nome"]:
        for n in wb.sheetnames:
            if n.strip().lower() in ("planilha1", "sheet1", "dados"):
                ws = wb[n]
                col_c = col_t = None
                header_row = None
                for row in ws.iter_rows(max_row=5, values_only=True):
                    cells = [str(c).strip().upper() if c is not None else "" for c in row]
                    if "CLIENTE" in cells:
                        col_c = cells.index("CLIENTE")
                        col_t = cells.index("CONTATO") if "CONTATO" in cells else None
                        header_row = True
                        break
                if header_row:
                    for data_row in ws.iter_rows(min_row=3, max_row=10, values_only=True):
                        if data_row[col_c]:
                            resultado["nome"] = str(data_row[col_c]).strip()
                            if col_t is not None and data_row[col_t]:
                                resultado["telefone"] = str(data_row[col_t]).strip()
                            break
                break

    wb.close()
    return resultado


def normalizar_telefone(tel):
    return re.sub(r"\D", "", tel)


def normalizar_nome(nome):
    return " ".join(nome.strip().title().split())


def gerar_crm(pasta):
    arquivos = listar_xlsx(pasta)
    if not arquivos:
        print(f"Nenhum arquivo .xlsx encontrado em: {pasta}")
        return

    print(f"\n{'='*60}")
    print(f"  {len(arquivos)} arquivo(s) encontrado(s)")
    print(f"{'='*60}\n")

    registros = []

    for i, caminho in enumerate(arquivos, 1):
        nome_arquivo = os.path.basename(caminho)
        label = nome_arquivo[:52]
        print(f"[{i:>4}/{len(arquivos)}] {label:<52}", end=" ", flush=True)

        dados = extrair_contato(caminho)

        if dados.get("erro"):
            print(f"ERRO: {dados['erro']}")
            continue

        nome = normalizar_nome(dados["nome"])
        telefone = normalizar_telefone(dados["telefone"])

        if nome or telefone:
            registros.append({
                "Nome": nome,
                "Telefone": telefone,
                "Arquivo_Origem": nome_arquivo
            })
            print(f"OK  {nome or '(sem nome)'} | {telefone or '(sem tel)'}")
        else:
            print("--  nao identificado")

    if not registros:
        print("\nNenhum contato extraido.")
        return

    df = pd.DataFrame(registros)

    # Deduplicacao por telefone
    df_com = df[df["Telefone"] != ""].copy()
    df_sem = df[df["Telefone"] == ""].copy()

    def agrupar(dframe, key):
        return (
            dframe.groupby(key, sort=False)
            .agg(
                Nome=("Nome", "first"),
                Telefone=("Telefone", "first"),
                Qtd_Pedidos=("Arquivo_Origem", "count"),
                Arquivos_Origem=("Arquivo_Origem", lambda x: "; ".join(x.unique()))
            )
            .reset_index(drop=True)
        )

    partes = []
    if not df_com.empty:
        partes.append(agrupar(df_com, "Telefone"))
    if not df_sem.empty:
        df_sem2 = df_sem.assign(chave=df_sem["Nome"].str.lower())
        partes.append(agrupar(df_sem2, "chave"))

    df_final = pd.concat(partes, ignore_index=True).sort_values("Nome").reset_index(drop=True)

    # Exportar
    saida = os.path.join(os.path.dirname(os.path.abspath(__file__)), OUTPUT_FILE)

    wb_out = Workbook()
    ws_out = wb_out.active
    ws_out.title = "CRM"

    colunas = ["Nome", "Telefone", "Qtd_Pedidos"]
    larguras = [35, 18, 14]

    header_fill = PatternFill("solid", start_color="1F4E79")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    alt_fill = PatternFill("solid", start_color="EBF3FB")

    for col_idx, (col, larg) in enumerate(zip(colunas, larguras), 1):
        cell = ws_out.cell(row=1, column=col_idx, value=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws_out.column_dimensions[get_column_letter(col_idx)].width = larg
    ws_out.row_dimensions[1].height = 22

    for row_idx, row in df_final.iterrows():
        excel_row = row_idx + 2
        for col_idx, col in enumerate(colunas, 1):
            cell = ws_out.cell(row=excel_row, column=col_idx, value=row[col])
            cell.alignment = Alignment(vertical="center", wrap_text=(col == "Arquivos_Origem"))
            if row_idx % 2 == 0:
                cell.fill = alt_fill

    ws_out.freeze_panes = "A2"
    wb_out.save(saida)

    removidas = len(registros) - len(df_final)
    print(f"\n{'='*60}")
    print(f"  {len(df_final)} contato(s) unicos gerados")
    print(f"  {removidas} duplicata(s) removida(s)")
    print(f"  Arquivo salvo em: {saida}")
    print(f"{'='*60}\n")


if __name__ == "__main__":
    if len(sys.argv) > 1:
        pasta_pedidos = sys.argv[1]
    else:
        pasta_pedidos = input("Caminho da pasta com os pedidos:\n> ").strip().strip('"')

    if not os.path.isdir(pasta_pedidos):
        print(f"Pasta nao encontrada: {pasta_pedidos}")
        sys.exit(1)

    gerar_crm(pasta_pedidos)
